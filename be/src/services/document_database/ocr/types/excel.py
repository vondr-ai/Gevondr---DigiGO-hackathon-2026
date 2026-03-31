from __future__ import annotations

import asyncio
import csv

from markitdown import MarkItDown

from src.services.document_database.ocr.doc_router import ProcessDecision
from src.services.document_database.ocr.page_format import format_document_with_pages
from src.services.document_database.ocr.types.interface import FileReader


class ExcelReader(FileReader):
    async def read(
        self,
        filename: str,
        path: str,
        fate: ProcessDecision,
    ) -> str:
        if fate == ProcessDecision.OCR:
            # Excel files don't require OCR
            raise ValueError("OCR is not applicable for Excel files")

        file_path = self._resolve_file_path(path, filename)
        suffix = file_path.suffix.lower()

        if suffix == ".csv":
            try:
                page_text = await asyncio.to_thread(
                    self._extract_csv_page_text_sync, str(file_path)
                )
                return format_document_with_pages(filename, [page_text])
            except Exception:
                pass

        try:
            page_texts = await asyncio.to_thread(
                self._extract_excel_pages_sync, str(file_path)
            )
            return format_document_with_pages(filename, page_texts)
        except Exception:
            md = MarkItDown(enable_plugins=False)  # Set to True to enable plugins
            result = await asyncio.to_thread(md.convert, str(file_path))
            return format_document_with_pages(filename, [result.text_content])

    def _extract_excel_pages_sync(self, file_path: str) -> list[str]:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        page_texts: list[str] = []
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                lines = [f"Sheet: {sheet_name}"]
                for row in sheet.iter_rows(values_only=True):
                    row_values = [
                        str(value).strip()
                        for value in row
                        if value is not None and str(value).strip()
                    ]
                    if row_values:
                        lines.append(" | ".join(row_values))
                page_texts.append("\n".join(lines).strip())
        finally:
            workbook.close()

        return page_texts or [""]

    def _extract_csv_page_text_sync(self, file_path: str) -> str:
        lines = ["Sheet: csv"]
        with open(file_path, "r", encoding="utf-8", errors="ignore") as handle:
            reader = csv.reader(handle)
            for row in reader:
                row_values = [value.strip() for value in row if value and value.strip()]
                if row_values:
                    lines.append(" | ".join(row_values))
        return "\n".join(lines).strip()

    def get_page_count(self, filename: str, path: str) -> int:
        """
        Count sheets in an Excel file (each sheet = 1 page).

        Uses openpyxl to count the number of sheets in the workbook.
        Falls back to 1 if openpyxl is not installed or if the file cannot be read.
        """
        file_path = self._resolve_file_path(path, filename)
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            try:
                return len(wb.sheetnames)
            finally:
                wb.close()
        except ImportError:
            # If openpyxl not installed, return 1
            return 1
        except Exception:
            return 1
