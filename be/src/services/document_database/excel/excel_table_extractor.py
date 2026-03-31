from __future__ import annotations

import csv
import logging
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence
from uuid import UUID

from markitdown import MarkItDown
from openpyxl import load_workbook

from src.database.postgres.excel.excel_table_repo import (
    ExcelTableMetadata,
    ExcelTableRepository,
    get_excel_table_repo,
)
from src.services.document_database.ocr.page_format import format_document_with_pages

logger = logging.getLogger(__name__)


OPENPYXL_EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
TABULAR_EXCEL_EXTENSIONS = {".xls", ".xlsb"}
EXCEL_EXTENSIONS = OPENPYXL_EXCEL_EXTENSIONS | TABULAR_EXCEL_EXTENSIONS
CSV_EXTENSIONS = {".csv"}
HEADER_SCAN_ROWS = 3
MIN_TABLE_ROWS = 50
MIN_COLUMNS = 2
CONSISTENCY_THRESHOLD = 0.7
MAX_NON_TABLE_LINES = 200


@dataclass
class ExcelDetectedTable:
    sheet_name: str
    table_name: str
    column_definitions: list[dict[str, str]]
    rows: list[list[object | None]]

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def column_count(self) -> int:
        return len(self.column_definitions)


@dataclass
class ExcelExtractionResult:
    full_text: str
    tables: list[ExcelDetectedTable]


class ExcelTableExtractor:
    def __init__(
        self,
        repo: ExcelTableRepository | None = None,
    ) -> None:
        self.repo = repo or get_excel_table_repo()

    def extract_and_store(
        self,
        path: str,
        filename: str,
        document_id: UUID,
        project_omgeving_id: UUID | None,
    ) -> ExcelExtractionResult:
        file_path = Path(path)
        extension = file_path.suffix.lower()
        if extension in CSV_EXTENSIONS:
            result = self._extract_csv(path, filename, document_id, project_omgeving_id)
            return result
        if extension in EXCEL_EXTENSIONS:
            result = self._extract_excel(
                path,
                filename,
                document_id,
                project_omgeving_id,
                extension=extension,
            )
            return result

        # Unsupported excel variant; fall back to MarkItDown
        content = self._safe_markitdown_text(path, filename)
        return ExcelExtractionResult(
            full_text=format_document_with_pages(filename, [content]),
            tables=[],
        )

    def _extract_excel(
        self,
        path: str,
        filename: str,
        document_id: UUID,
        project_omgeving_id: UUID | None,
        extension: str | None = None,
    ) -> ExcelExtractionResult:
        file_extension = (extension or Path(path).suffix.lower()).lower()
        if file_extension in TABULAR_EXCEL_EXTENSIONS:
            return self._extract_excel_with_tabular_reader(
                path=path,
                filename=filename,
                document_id=document_id,
                project_omgeving_id=project_omgeving_id,
                extension=file_extension,
            )
        return self._extract_excel_with_openpyxl(
            path=path,
            filename=filename,
            document_id=document_id,
            project_omgeving_id=project_omgeving_id,
        )

    def _extract_excel_with_openpyxl(
        self,
        path: str,
        filename: str,
        document_id: UUID,
        project_omgeving_id: UUID | None,
    ) -> ExcelExtractionResult:
        try:
            workbook = load_workbook(path, data_only=True, read_only=False)
        except Exception:
            content = self._safe_markitdown_text(path, filename)
            return ExcelExtractionResult(
                full_text=format_document_with_pages(filename, [content]),
                tables=[],
            )

        tables: list[ExcelDetectedTable] = []
        non_table_parts: list[str] = []
        try:
            for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
                sheet = workbook[sheet_name]
                sheet_rows = [
                    [cell for cell in row] for row in sheet.iter_rows(values_only=True)
                ]
                merged_cells = getattr(sheet, "merged_cells", None)
                merged_ranges = (
                    getattr(merged_cells, "ranges", []) if merged_cells else []
                )
                if merged_ranges:
                    non_table_parts.append(
                        self._rows_to_text(sheet_rows, sheet_name, None)
                    )
                    continue

                table, table_bounds = self._extract_sheet_table(
                    sheet_rows=sheet_rows,
                    sheet_name=sheet_name,
                    sheet_index=sheet_index,
                    document_id=document_id,
                )
                if table is None or table_bounds is None:
                    non_table_parts.append(
                        self._rows_to_text(sheet_rows, sheet_name, None)
                    )
                    continue

                tables.append(table)
                non_table_parts.append(
                    self._rows_to_text(sheet_rows, sheet_name, table_bounds)
                )
        finally:
            workbook.close()

        if tables:
            self._store_tables(
                document_id,
                project_omgeving_id,
                filename,
                tables,
            )
            full_text = self._build_full_text(filename, tables, non_table_parts)
            return ExcelExtractionResult(full_text=full_text, tables=tables)

        content = self._safe_markitdown_text(path, filename)
        return ExcelExtractionResult(
            full_text=format_document_with_pages(filename, [content]),
            tables=[],
        )

    def _extract_excel_with_tabular_reader(
        self,
        path: str,
        filename: str,
        document_id: UUID,
        project_omgeving_id: UUID | None,
        extension: str,
    ) -> ExcelExtractionResult:
        try:
            import pandas as pd
        except Exception:
            content = self._safe_markitdown_text(path, filename)
            return ExcelExtractionResult(
                full_text=format_document_with_pages(filename, [content]),
                tables=[],
            )

        engine_by_extension = {
            ".xls": "xlrd",
            ".xlsb": "pyxlsb",
        }
        engine = engine_by_extension.get(extension)

        try:
            workbook = pd.ExcelFile(path, engine=engine)
        except Exception:
            content = self._safe_markitdown_text(path, filename)
            return ExcelExtractionResult(
                full_text=format_document_with_pages(filename, [content]),
                tables=[],
            )

        tables: list[ExcelDetectedTable] = []
        non_table_parts: list[str] = []
        try:
            for sheet_index, sheet_name in enumerate(workbook.sheet_names, start=1):
                raw_frame = workbook.parse(
                    sheet_name=sheet_name, header=None, dtype=object
                )
                assert isinstance(raw_frame, pd.DataFrame)
                sheet_rows = [
                    [self._normalize_tabular_value(value) for value in row]
                    for row in raw_frame.itertuples(index=False, name=None)
                ]
                table, table_bounds = self._extract_sheet_table(
                    sheet_rows=sheet_rows,
                    sheet_name=sheet_name,
                    sheet_index=sheet_index,
                    document_id=document_id,
                )
                if table is None or table_bounds is None:
                    non_table_parts.append(
                        self._rows_to_text(sheet_rows, sheet_name, None)
                    )
                    continue

                tables.append(table)
                non_table_parts.append(
                    self._rows_to_text(sheet_rows, sheet_name, table_bounds)
                )
        finally:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()

        if tables:
            self._store_tables(
                document_id,
                project_omgeving_id,
                filename,
                tables,
            )
            full_text = self._build_full_text(filename, tables, non_table_parts)
            return ExcelExtractionResult(full_text=full_text, tables=tables)

        content = self._safe_markitdown_text(path, filename)
        return ExcelExtractionResult(
            full_text=format_document_with_pages(filename, [content]),
            tables=[],
        )

    def _extract_csv(
        self,
        path: str,
        filename: str,
        document_id: UUID,
        project_omgeving_id: UUID | None,
    ) -> ExcelExtractionResult:
        rows: list[list[object | None]] = []
        if not self._is_valid_csv(path):
            content = self._safe_markitdown_text(path, filename)
            return ExcelExtractionResult(
                full_text=format_document_with_pages(filename, [content]),
                tables=[],
            )

        with open(path, "r", encoding="utf-8", errors="ignore") as handle:
            reader = csv.reader(handle)
            for row in reader:
                rows.append([self._parse_csv_value(cell) for cell in row])

        detection = self._detect_table_from_rows(rows)
        if detection is None:
            content = self._safe_markitdown_text(path, filename)
            return ExcelExtractionResult(
                full_text=format_document_with_pages(filename, [content]),
                tables=[],
            )

        header_row_idx, col_start, col_end, data_rows = detection
        headers = rows[header_row_idx][col_start:col_end]
        sanitized_headers = self._sanitize_headers(headers)
        column_defs = self._infer_column_definitions(
            headers, sanitized_headers, data_rows
        )
        if column_defs is None:
            content = self._safe_markitdown_text(path, filename)
            return ExcelExtractionResult(
                full_text=format_document_with_pages(filename, [content]),
                tables=[],
            )

        table_name = self._build_table_name(
            document_id=document_id,
            sheet_name="csv",
            sheet_index=1,
        )
        table = ExcelDetectedTable(
            sheet_name="csv",
            table_name=table_name,
            column_definitions=column_defs,
            rows=data_rows,
        )
        self._store_tables(document_id, project_omgeving_id, filename, [table])
        full_text = self._build_full_text(filename, [table], [])
        return ExcelExtractionResult(full_text=full_text, tables=[table])

    def _safe_markitdown_text(self, path: str, filename: str) -> str:
        try:
            md = MarkItDown(enable_plugins=False)
            return md.convert(str(path)).text_content
        except Exception as exc:
            logger.warning(
                "MarkItDown fallback failed for %s (%s): %s",
                filename,
                path,
                exc,
            )
            return "[Text extraction failed for this file.]"

    def _store_tables(
        self,
        document_id: UUID,
        project_omgeving_id: UUID | None,
        filename: str,
        tables: Iterable[ExcelDetectedTable],
    ) -> None:
        metadata_list: list[ExcelTableMetadata] = []
        rows_by_table: dict[str, list[list[object | None]]] = {}
        for table in tables:
            metadata_list.append(
                ExcelTableMetadata(
                    document_id=document_id,
                    project_omgeving_id=project_omgeving_id,
                    filename=filename,
                    sheet_name=table.sheet_name,
                    table_name=table.table_name,
                    column_definitions=table.column_definitions,
                    row_count=table.row_count,
                    column_count=table.column_count,
                )
            )
            rows_by_table[table.table_name] = table.rows

        self.repo.upsert_document_tables(
            document_id=document_id,
            project_omgeving_id=project_omgeving_id,
            filename=filename,
            tables=metadata_list,
            rows_by_table=rows_by_table,
        )

    def _extract_sheet_table(
        self,
        *,
        sheet_rows: Sequence[Sequence[object | None]],
        sheet_name: str,
        sheet_index: int,
        document_id: UUID,
    ) -> tuple[ExcelDetectedTable | None, tuple[int, int, int, int] | None]:
        detection = self._detect_table_from_rows(sheet_rows)
        if detection is None:
            return None, None

        header_row_idx, col_start, col_end, data_rows = detection
        headers = sheet_rows[header_row_idx][col_start:col_end]
        sanitized_headers = self._sanitize_headers(headers)
        column_defs = self._infer_column_definitions(
            headers, sanitized_headers, data_rows
        )
        if column_defs is None:
            return None, None

        table = ExcelDetectedTable(
            sheet_name=sheet_name,
            table_name=self._build_table_name(
                document_id=document_id,
                sheet_name=sheet_name,
                sheet_index=sheet_index,
            ),
            column_definitions=column_defs,
            rows=data_rows,
        )
        table_bounds = (
            header_row_idx + 1,
            col_start + 1,
            col_end,
            header_row_idx + 1 + len(data_rows),
        )
        return table, table_bounds

    def _detect_table_from_rows(
        self, rows: Sequence[Sequence[object | None]]
    ) -> tuple[int, int, int, list[list[object | None]]] | None:
        if not rows:
            return None

        max_scan = min(HEADER_SCAN_ROWS, len(rows))
        for header_idx in range(max_scan):
            header_row = rows[header_idx]
            col_start, col_end = self._find_header_bounds(header_row)
            if col_start is None or col_end is None:
                continue
            headers = header_row[col_start:col_end]
            if len(headers) < MIN_COLUMNS:
                continue
            data_rows = self._collect_data_rows(
                rows[header_idx + 1 :], col_start, col_end
            )
            if len(data_rows) < MIN_TABLE_ROWS:
                continue
            if not self._columns_consistent(data_rows):
                continue
            return header_idx, col_start, col_end, data_rows

        return None

    def _find_header_bounds(
        self, row: Sequence[object | None]
    ) -> tuple[int | None, int | None]:
        normalized = [
            str(cell).strip() if cell is not None and str(cell).strip() else ""
            for cell in row
        ]
        if not any(normalized):
            return None, None
        try:
            start = next(idx for idx, value in enumerate(normalized) if value)
        except StopIteration:
            return None, None
        end = start
        for idx in range(start, len(normalized)):
            if normalized[idx]:
                end = idx
            else:
                break
        return start, end + 1

    def _collect_data_rows(
        self,
        rows: Sequence[Sequence[object | None]],
        col_start: int,
        col_end: int,
    ) -> list[list[object | None]]:
        data_rows: list[list[object | None]] = []
        empty_gap = 0
        expected_len = col_end - col_start
        for row in rows:
            row_slice = list(row[col_start:col_end])
            if len(row_slice) < expected_len:
                row_slice = row_slice + [None] * (expected_len - len(row_slice))
            if self._row_is_empty(row_slice):
                empty_gap += 1
                if empty_gap >= 5:
                    break
                continue
            empty_gap = 0
            data_rows.append([self._normalize_value(value) for value in row_slice])
        return data_rows

    def _row_is_empty(self, row: Sequence[object | None]) -> bool:
        for value in row:
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return False
        return True

    def _columns_consistent(self, rows: Sequence[Sequence[object | None]]) -> bool:
        if not rows:
            return False
        columns = list(zip(*rows))
        for column in columns:
            if not self._column_consistency_ok(column):
                return False
        return True

    def _column_consistency_ok(self, values: Iterable[object | None]) -> bool:
        cleaned = [value for value in values if value is not None]
        if not cleaned:
            return True
        types = [self._classify_value(value) for value in cleaned]
        counts = Counter(types)
        total = sum(counts.values())

        numeric_count = counts.get("int", 0) + counts.get("float", 0)
        if numeric_count / total >= CONSISTENCY_THRESHOLD:
            return True
        if counts.get("datetime", 0) / total >= CONSISTENCY_THRESHOLD:
            return True
        if counts.get("bool", 0) / total >= CONSISTENCY_THRESHOLD:
            return True
        if counts.get("text", 0) / total >= CONSISTENCY_THRESHOLD:
            return True
        return False

    def _infer_column_definitions(
        self,
        raw_headers: Sequence[object | None],
        headers: list[str],
        rows: Sequence[Sequence[object | None]],
    ) -> list[dict[str, str]] | None:
        if not rows:
            return None
        column_defs: list[dict[str, str]] = []
        columns = list(zip(*rows))
        for raw_header, header, column_values in zip(raw_headers, headers, columns):
            label = str(raw_header).strip() if raw_header is not None else ""
            column_type = self._infer_column_type(column_values)
            if column_type is None:
                return None
            column_defs.append(
                {"name": header, "type": column_type, "label": label or header}
            )
        return column_defs

    def _infer_column_type(self, values: Iterable[object | None]) -> str | None:
        cleaned = [value for value in values if value is not None]
        if not cleaned:
            return "TEXT"
        types = [self._classify_value(value) for value in cleaned]
        counts = Counter(types)
        total = sum(counts.values())

        numeric_count = counts.get("int", 0) + counts.get("float", 0)
        if numeric_count / total >= CONSISTENCY_THRESHOLD:
            if counts.get("float", 0) > 0:
                return "DOUBLE PRECISION"
            return "BIGINT"
        if counts.get("datetime", 0) / total >= CONSISTENCY_THRESHOLD:
            return "TIMESTAMPTZ"
        if counts.get("bool", 0) / total >= CONSISTENCY_THRESHOLD:
            return "BOOLEAN"
        if counts.get("text", 0) / total >= CONSISTENCY_THRESHOLD:
            return "TEXT"
        return None

    def _classify_value(self, value: object) -> str:
        if isinstance(value, bool):
            return "bool"
        if isinstance(value, int) and not isinstance(value, bool):
            return "int"
        if isinstance(value, float):
            return "float"
        if isinstance(value, datetime):
            return "datetime"
        if isinstance(value, date):
            return "datetime"
        return "text"

    def _build_table_name(
        self,
        document_id: UUID,
        sheet_name: str,
        sheet_index: int,
    ) -> str:
        doc_token = document_id.hex[:8]
        sheet_token = self._sanitize_identifier(sheet_name)[:24] or "sheet"
        return f"excel_{doc_token}_{sheet_index}_{sheet_token}"

    def _sanitize_headers(self, headers: Sequence[object | None]) -> list[str]:
        cleaned: list[str] = []
        seen: dict[str, int] = {}
        for idx, header in enumerate(headers, start=1):
            name = self._sanitize_identifier(str(header) if header is not None else "")
            if not name:
                name = f"column_{idx}"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 1
            cleaned.append(name)
        return cleaned

    def _sanitize_identifier(self, value: str) -> str:
        value = value.strip().lower()
        value = re.sub(r"\s+", "_", value)
        value = re.sub(r"[^a-z0-9_]", "", value)
        return value

    def _normalize_value(self, value: object | None) -> object | None:
        if value is None:
            return None
        if isinstance(value, str):
            stripped = value.strip()
            return stripped if stripped else None
        return value

    def _normalize_tabular_value(self, value: object | None) -> object | None:
        if value is None:
            return None

        to_python = getattr(value, "to_pydatetime", None)
        if callable(to_python):
            value = to_python()

        try:
            is_nan = value != value
        except Exception:
            is_nan = False
        if is_nan is True:
            return None

        return self._normalize_value(value)

    def _rows_to_text(
        self,
        rows: Sequence[Sequence[object | None]],
        sheet_name: str,
        table_bounds: tuple[int, int, int, int] | None,
    ) -> str:
        lines = [f"Sheet: {sheet_name}"]
        line_count = 0
        for row_idx, row in enumerate(rows, start=1):
            row_values: list[str] = []
            for col_idx, value in enumerate(row, start=1):
                if self._cell_in_table_bounds(row_idx, col_idx, table_bounds):
                    continue
                if value is None:
                    continue
                if isinstance(value, str) and not value.strip():
                    continue
                row_values.append(str(value))
            if row_values:
                lines.append(" | ".join(row_values))
                line_count += 1
                if line_count >= MAX_NON_TABLE_LINES:
                    lines.append("[Non-table text truncated]")
                    break
        return "\n".join(lines)

    def _cell_in_table_bounds(
        self,
        row_idx: int,
        col_idx: int,
        table_bounds: tuple[int, int, int, int] | None,
    ) -> bool:
        if not table_bounds:
            return False
        start_row, start_col, end_col, end_row = table_bounds
        return (
            row_idx >= start_row
            and row_idx <= end_row
            and col_idx >= start_col
            and col_idx <= end_col
        )

    def _build_full_text(
        self,
        filename: str,
        tables: list[ExcelDetectedTable],
        non_table_parts: list[str],
    ) -> str:
        table_lines_by_sheet: dict[str, list[str]] = {}
        for table in tables:
            table_lines_by_sheet.setdefault(table.sheet_name, []).append(
                self._build_table_summary_line(table)
            )

        page_texts: list[str] = []
        for part in non_table_parts:
            sheet_name = self._extract_sheet_name(part)
            lines: list[str] = []
            lines.append("Detected tables:")
            if sheet_name and table_lines_by_sheet.get(sheet_name):
                lines.extend(table_lines_by_sheet[sheet_name])
            else:
                lines.append("- None detected on this sheet.")
            lines.append(part.strip())
            page_texts.append("\n".join(lines).strip())

        if not page_texts:
            lines = ["Detected tables:"]
            if tables:
                lines.extend(self._build_table_summary_line(table) for table in tables)
            else:
                lines.append("[No extractable sheet content.]")
            page_texts = ["\n".join(lines)]

        return format_document_with_pages(filename, page_texts)

    def _build_table_summary_line(self, table: ExcelDetectedTable) -> str:
        columns = ", ".join(
            f"{col['label']} -> {col['name']} ({col['type']})"
            for col in table.column_definitions
        )
        return (
            f"- Sheet '{table.sheet_name}' -> table '{table.table_name}' "
            f"({table.row_count} rows, {table.column_count} columns). Columns: {columns}"
        )

    def _extract_sheet_name(self, sheet_text: str) -> str | None:
        lines = sheet_text.splitlines()
        if not lines:
            return None
        first_line = lines[0].strip()
        if not first_line.startswith("Sheet: "):
            return None
        return first_line.replace("Sheet: ", "", 1).strip() or None

    def _parse_csv_value(self, value: str) -> object | None:
        stripped = value.strip()
        if not stripped:
            return None
        lowered = stripped.lower()
        if lowered in {"true", "false", "yes", "no"}:
            return lowered in {"true", "yes"}
        if re.fullmatch(r"-?\d+", stripped):
            try:
                return int(stripped)
            except ValueError:
                return stripped
        if re.fullmatch(r"-?\d+\.\d+", stripped):
            try:
                return float(stripped)
            except ValueError:
                return stripped
        return stripped

    def _is_valid_csv(self, path: str) -> bool:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as handle:
                sample = handle.read(4096)
                if not sample.strip():
                    return False
                dialect = csv.Sniffer().sniff(sample)
                handle.seek(0)
                reader = csv.reader(handle, dialect)
                row_lengths = []
                for idx, row in enumerate(reader):
                    if idx >= 10:
                        break
                    row_lengths.append(len(row))
                if not row_lengths:
                    return False
                return max(row_lengths) == min(row_lengths)
        except Exception:
            return False
